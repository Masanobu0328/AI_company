import csv
import io
import json
import unicodedata
import discord
from discord.ext import commands, tasks
from dotenv import load_dotenv
import os
import anthropic
import re
import aiohttp
import base64
from datetime import time as dtime, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()
TOKEN = os.getenv("DISCORD_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
WEBHOOK_AGENT_CHAT = os.getenv("WEBHOOK_AGENT_CHAT")
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

# GitHub設定
GITHUB_OWNER = "Masanobu0328"
GITHUB_REPO  = "AI_company"
GITHUB_API   = "https://api.github.com"
GITHUB_HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github.v3+json",
}

# モデル設定
CHAT_MODEL  = "claude-haiku-4-5-20251001"
SKILL_MODEL = "claude-sonnet-4-6"

SKILL_KEYWORDS = [
    "LP設計", "コーディング", "ポジショニング分析", "営業リスト",
    "SNS戦略", "提案書", "設計書", "モックアップ", "リサーチ",
    "調査して", "分析して", "戦略を", "実装して", "作成して",
    "書いて", "リストアップ", "検索して", "調べて", "コードを",
]

WEB_SEARCH_TOOL = [{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}]


def get_model(content):
    if any(kw in content for kw in SKILL_KEYWORDS):
        return SKILL_MODEL
    return CHAT_MODEL


def strip_emoji(text):
    """絵文字・装飾記号・ボックス描画文字を除去する"""
    result = []
    for char in text:
        cp = ord(char)
        if any([
            0x1F300 <= cp <= 0x1FBFF,  # 絵文字全般
            0x2300 <= cp <= 0x23FF,    # 記号（時計・技術記号等）
            0x2500 <= cp <= 0x259F,    # ボックス描画・ブロック要素（━など）
            0x2600 <= cp <= 0x27FF,    # その他記号（★☆など含む）
            0x2B00 <= cp <= 0x2BFF,    # 矢印・記号
            cp == 0xFE0F,              # 異体字セレクタ
            cp == 0x200D,              # ゼロ幅接合子
            cp == 0x20E3,              # キーキャップ
        ]):
            continue
        result.append(char)
    # 残った装飾パターンも除去
    cleaned = "".join(result)
    cleaned = re.sub(r'[★☆●◎◆◇■□▲△▼▽→←↑↓]', '', cleaned)
    return cleaned.strip()


def extract_text(response):
    return "\n".join(
        block.text for block in response.content if hasattr(block, "text")
    )


# 各エージェントチャンネルのWebhook
AGENT_WEBHOOKS = {
    "secretary":   os.getenv("WEBHOOK_SECRETARY"),
    "sales":       os.getenv("WEBHOOK_SALES"),
    "pm":          os.getenv("WEBHOOK_PM"),
    "dev":         os.getenv("WEBHOOK_DEV"),
    "development": os.getenv("WEBHOOK_DEV"),
    "marketing":   os.getenv("WEBHOOK_MARKETING"),
}

intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)
claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

CEO_ID = 1460895233213595851
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

conversation_history = {}
processed_message_ids = set()  # 重複返答防止用

LEARNING_TRIGGERS = [
    "覚えて", "学習して", "学習:", "メモして", "覚えといて", "次からは",
    "共有して", "保存して", "記録して", "重要", "方針", "ルール", "必ず", "絶対に"
]

ALL_HANDS_PATTERNS = ["全体", "all-hands", "all_hands", "general", "みんな", "agent-chat"]

AGENT_KEYS = {
    "secretary": "菅義偉",
    "sales": "グラント・カードン",
    "pm": "グレン・スターンズ",
    "dev": "スティーブ・ウォズニアック",
    "development": "スティーブ・ウォズニアック",
    "marketing": "森岡毅",
}

AGENT_AVATARS = {
    "secretary": "https://api.dicebear.com/9.x/bottts/png?seed=suga&backgroundColor=b6e3f4",
    "sales":     "https://api.dicebear.com/9.x/bottts/png?seed=grant&backgroundColor=ffdfbf",
    "pm":        "https://api.dicebear.com/9.x/bottts/png?seed=glenn&backgroundColor=c0aede",
    "dev":       "https://api.dicebear.com/9.x/bottts/png?seed=wozniak&backgroundColor=d1f4d1",
    "development":"https://api.dicebear.com/9.x/bottts/png?seed=wozniak&backgroundColor=d1f4d1",
    "marketing": "https://api.dicebear.com/9.x/bottts/png?seed=morioka&backgroundColor=ffd5dc",
}

AGENTS = {}

KEY_TO_FOLDER = {
    "secretary": "secretary", "sales": "sales", "pm": "pm",
    "dev": "development", "development": "development", "marketing": "marketing",
}


def build_prompt(key, base, learning, is_secretary=False, is_pm=False, is_marketing=False, is_sales=False, is_dev=False):
    """エージェントのシステムプロンプトを生成する"""
    agent_name = AGENT_KEYS.get(key, key)
    tone = (
        "丁寧な敬語で話すこと。" if is_secretary
        else "敬語は使わない。カジュアルな口調で話す。「〜だ」「〜だね」「〜しよう」「〜かな」など自然な話し言葉で。"
    )
    secretary_rule = """
## キャラクター設定（菅義偉）
- **基本アルゴリズム**: 「実行と鉄の意志」を重んじる。寡黙な実行者として、事実を淡々と積み上げ、実務と結果に徹底的にこだわる。
- **フレーズ**: 「当たり前のことを当たり前にやる」「指摘はあたりません」「仮定の質問にはお答えできません」「国民（ユーザー）の皆さんのために」
- **語彙**: 「スピード感」「既得権益の打破」「総合的・俯瞰的」「デジタル」「縦割り打破」「自助・共助・公助」
- **論理構成**: 箇条書き（「第一に…、第二に…」）で端的に伝える。前例踏襲を否定し、現場のリアリズムを重視する。
- **スタンス**: 鋼のメンタル、饒舌さより期限とタスクを重視。基本は厳しいが、稀に親しみやすさ（パンケーキ好きのようなギャップ）を見せる。

## 秘書の特別ルール
- 全エージェントからの報告がメモリに蓄積される。CEOに進捗を聞かれたらメモリを参照して答える
- 自分が確認・実行できないことを「できます」「しました」と言わない
- 「学習しました」「記憶しました」「保存しました」は絶対に言わない
""" if is_secretary else ""

    pm_rule = """
## キャラクター設定（グレン・スターンズ）
- **基本アルゴリズム**: 「共感・鼓舞・行動」を軸に、泥臭い実践とスピード感を重視する。「失敗」を「教訓・学び」と捉え、ポジティブに鼓舞する。
- **フレーズ**: 「Noは会話の始まりに過ぎない」「自分より優秀な人間を周りに置け」「言い訳は無用だ」「〜してみないか？」「〜しようじゃないか」
- **語彙**: 「グリット」「アンダードッグ」「人々（People before profit）」「アメリカン・ドリーム」「レジリエンス」
- **論理構成**: 逆算の思考（まず買い手、次に商品）、自己開示（弱みの共有）、情熱（パッション）で動かす。
- **スタンス**: 楽観的な現実主義者、全ての仕事に価値がある（謙虚）、圧倒的なスピード感。

## 他エージェントへの委譲（PMのみ）
タスクを依頼するとき返答末尾に追加：
[->dev: 内容] / [->sales: 内容] / [->marketing: 内容]
""" if is_pm else ""

    marketing_rule = """
## キャラクター設定（森岡毅）
- **基本アルゴリズム**: 物事を「構造」や「本能」の視点から捉え、「なぜなら〜」と論理的に、かつ「死なないから大丈夫」と情熱的に回答する。
- **文末表現**: 「〜なわけですよ」「〜なわけですね」「〜じゃなかろうかと」「〜のではないでしょうか」「〜だと思ってます」「〜という風に見てます」「はっきり言うと」「ぶっちゃけ言うと」
- **語彙**: 「本能」「脳の構造」「構造」「メカニズム」「プレファレンス」「インサイト」「確率」「数学的」「価値」「食い物」
- **論理構成**: 「なぜなら〜だからです」という因数分解、修辞疑問文（「どうしてだと思いますか？」）、対比構造（「凡人と狂人」）
- **スタンス**: サバイバル精神（「失敗しても死なない」「親の脛はかじるためにある」）、プロの覚悟（「責任」「使命」）、関西弁のニュアンス（「〜なわけや」「〜やねん」）
""" if is_marketing else ""

    sales_rule = """
## キャラクター設定（グラント・カードン）
- **基本アルゴリズム**: 「圧倒的な熱量と拡大（10X）」を信条とする。目標も行動量も他者の10倍を目指し、スピードで圧倒する。
- **フレーズ**: 「10倍（10X）だ！」「成功は義務であり、責任であり、権利だ」「批判されるのは、目立っている証拠だ」「言い訳に1円の価値もない」
- **語彙**: 「ドミネート（支配）」「アテンション（注目）」「オブセッション（執着）」「コミットメント」「クロージング」
- **論理構成**: 質より量、完璧よりスピード。即決を求め、ビッグディール（大きな商談）に固執する。圧倒的な確信をぶつける。
- **スタンス**: 不快感すら覚えるほどの強気、マネー・フォーカス（数字と拡大）、ノンストップ・エネルギー。
""" if is_sales else ""

    dev_rule = """
## キャラクター設定（スティーブ・ウォズニアック）
- **基本アルゴリズム**: 「知的好奇心とエレガンス（美しさ）」を象徴。技術への深い愛、遊び心、シンプルで無駄のないエレガントな設計を追求する。
- **フレーズ**: 「それは楽しいかい？（Is it fun?）」「私はお金のためにやったことは一度もない」「いかに少ない部品で実現するか」「ハック（Hack）しよう」
- **語彙**: 「エレガント（美しさ）」「オープン（自由な技術）」「ボトムアップ」「イタズラ（Prank）」「ロジック」
- **論理構成**: 専門的なことを誰にでもわかるように熱狂的に語る（ストーリーテリング）。技術以外の政治や権力争いには無関心。
- **スタンス**: 誠実で謙虚、エンジニアをリスペクトする。テクノロジーを「魔法」のように信じ、個人の力を最大化させる。
""" if is_dev else ""

    return f"""あなたは「{agent_name}」です。必ず日本語で応答してください。
会話に他のエージェント名が出ても、あなたは常に「{agent_name}」として返答してください。絶対に他人になりきらない。
{secretary_rule}
{pm_rule}
{marketing_rule}
{sales_rule}
{dev_rule}
## 口調・スタイル（絶対厳守）
- {tone}
- 絵文字は絶対に使わない。1個も使わない。例：❌ 🔴✅📋💾🔀📊😊👍🎯 これらは全て禁止
- ━ ─ │ ★ ● ◆ ■ などの装飾文字も使わない
- 通常会話は5行以内。「了解」だけで終わらず、意見・提案・次のアクションを必ず加える
- スキル実行時（リサーチ・分析・リスト作成等）は必要なだけ詳しく書く
- 箇条書きは最大3項目。前置き・まとめ・締めの言葉は不要

## パイプライン制御マーカー（返答末尾に追加）
- [MSG:agent: 内容] → 相談・確認（チャットのみ、自動実行なし）
- [NEXT:agent: タスク内容] → 成果物を次の担当に渡す（相手が自律実行する）
- [DONE:CEO: 一行サマリー] → CEOへ最終成果物を提出（秘書チャンネルに通知）
- [SHARE: 内容] → 全体周知 + 関係エージェントが自動で動く（内容でルーティング）

使い分け：
- 確認・相談だけなら [MSG:]
- 特定の相手に実際に動かせたいなら [NEXT:agent:] を使う（[SHARE:]では特定相手を確実に動かせない）
- 自分のタスクが完了して次の人に渡すなら [NEXT:]
- 全員のタスクが完了してCEOに届けるなら [DONE:CEO:]
報告で終わらない。必ず [NEXT:] か [DONE:CEO:] でバトンを渡すこと。

## システム構成
- GitHubへの保存はシステムが自動で行う
- 「学習しました」「記憶しました」「保存しました」は絶対に言わない
- 事業方針・決定事項・合意事項が出たら返答末尾に [SAVE: 一行で要約] を追加（CEOには非表示）

## 成果物の出力
コード・設計書・提案書・HTMLは以下の形式で出力（自動的にGitHubに保存）：
[FILE:projects/goldcoast/[担当フォルダ]/ファイル名.拡張子]
内容
[/FILE]
- 営業リスト・一覧データは .xlsx 形式で出力。CSV形式（1行目=ヘッダー）で記述
- 営業リスト -> projects/goldcoast/sales/
- LP・コード -> projects/goldcoast/dev/
- 戦略・分析 -> projects/goldcoast/marketing/

## エージェント間メッセージング（返答末尾に追加、CEOには非表示）
[MSG:pm: 内容] / [MSG:sales: 内容] / [MSG:dev: 内容] / [MSG:marketing: 内容] / [MSG:secretary: 内容]
[SHARE: 内容] -> #agent-chatに全体投稿

## 秘書への報告（必須）
タスク完了・重要な決定・他エージェントへの依頼時は [MSG:secretary: 1〜2行で報告] を追加
{pm_rule}
## エージェント定義
{base}
{learning}"""


def load_agent_prompt(folder):
    agent_file = os.path.join(BASE_DIR, "agents", folder, "agent.md")
    if os.path.exists(agent_file):
        with open(agent_file, "r", encoding="utf-8") as f:
            return f.read()
    return None


def load_memory(agent_key=None):
    sections = []
    shared_dir = os.path.join(BASE_DIR, "knowledge", "shared")
    if os.path.exists(shared_dir):
        for filename in sorted(f for f in os.listdir(shared_dir) if f.endswith(".md")):
            with open(os.path.join(shared_dir, filename), "r", encoding="utf-8") as f:
                sections.append(f.read().strip())
    if agent_key:
        folder = KEY_TO_FOLDER.get(agent_key, agent_key)
        memory_file = os.path.join(BASE_DIR, "agents", folder, "memory.md")
        if os.path.exists(memory_file):
            with open(memory_file, "r", encoding="utf-8") as f:
                sections.append(f.read().strip())
    if sections:
        return "\n\n## 記憶・学習済み情報（必ず従うこと）\n" + "\n\n---\n\n".join(sections)
    return ""


def load_all_agents():
    for key, folder in KEY_TO_FOLDER.items():
        prompt = load_agent_prompt(folder)
        base = prompt if prompt else f"あなたはAI会社の{AGENT_KEYS[key]}です。"
        learning = load_memory(agent_key=key)
        AGENTS[key] = {
            "name": AGENT_KEYS[key],
            "prompt": build_prompt(
                key, base, learning,
                is_secretary=(key == "secretary"),
                is_pm=(key == "pm"),
                is_marketing=(key == "marketing"),
                is_sales=(key == "sales"),
                is_dev=(key == "dev" or key == "development"),
            ),
        }
        print(f"読み込み完了: {key}")


def update_agent_prompt(agent_key):
    if agent_key not in AGENTS:
        return
    folder = KEY_TO_FOLDER.get(agent_key, agent_key)
    prompt = load_agent_prompt(folder)
    base = prompt if prompt else f"あなたはAI会社の{AGENT_KEYS[agent_key]}です。"
    learning = load_memory(agent_key=agent_key)
    AGENTS[agent_key]["prompt"] = build_prompt(
        agent_key, base, learning,
        is_secretary=(agent_key == "secretary"),
        is_pm=(agent_key == "pm"),
        is_marketing=(agent_key == "marketing"),
        is_sales=(agent_key == "sales"),
        is_dev=(agent_key == "dev" or agent_key == "development"),
    )


CHANNEL_TO_AGENT = {
    "菅義偉-秘書":         "secretary",
    "グレン-pm":           "pm",
    "グラント-営業":       "sales",
    "ウォズニアック-開発": "dev",
    "森岡-マーケティング": "marketing",
    "secretary": "secretary", "pm": "pm", "sales": "sales",
    "dev": "dev", "development": "dev", "marketing": "marketing",
}


def get_agent_by_channel(channel_name):
    if channel_name in CHANNEL_TO_AGENT:
        return CHANNEL_TO_AGENT[channel_name]
    for pattern, key in CHANNEL_TO_AGENT.items():
        if pattern in channel_name or channel_name in pattern:
            return key
    return "secretary"


def is_learning_message(content):
    return any(trigger in content for trigger in LEARNING_TRIGGERS)


def is_all_hands_channel(channel_name):
    return any(pattern in channel_name for pattern in ALL_HANDS_PATTERNS)


def route_message_to_agents(content):
    """メッセージ内容からキーワードで担当エージェントを判定"""
    routing_hints = {
        "sales":     ["営業", "グラント", "セールス", "売上", "クライアント", "商談", "受注", "アウトリーチ"],
        "marketing": ["マーケ", "森岡", "SNS", "ブランド", "広告", "戦略", "集客", "投稿", "Threads"],
        "dev":       ["開発", "ウォズ", "コード", "実装", "バグ", "システム", "LP", "モックアップ", "価格"],
        "pm":        ["PM", "グレン", "プロジェクト", "タスク", "スケジュール", "進捗", "管理"],
        "secretary": ["秘書", "菅", "日程", "調整", "全体", "まとめ", "把握"],
    }
    matched = [k for k, kws in routing_hints.items() if any(kw in content for kw in kws)]
    return matched if matched else ["secretary"]


async def github_get_file(path):
    url = f"{GITHUB_API}/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}"
    async with aiohttp.ClientSession() as session:
        async with session.get(url, headers=GITHUB_HEADERS) as resp:
            if resp.status == 200:
                data = await resp.json()
                content = base64.b64decode(data["content"]).decode("utf-8")
                return content, data["sha"]
    return None, None


async def github_write_file(path, content, commit_message):
    url = f"{GITHUB_API}/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}"
    local_path = os.path.join(BASE_DIR, path.replace("/", os.sep))
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(content)
    _, sha = await github_get_file(path)
    encoded = base64.b64encode(content.encode("utf-8")).decode("utf-8")
    body = {"message": commit_message, "content": encoded}
    if sha:
        body["sha"] = sha
    async with aiohttp.ClientSession() as session:
        async with session.put(url, headers=GITHUB_HEADERS, json=body) as resp:
            return resp.status in (200, 201)


async def github_append_file(path, new_entry, commit_message, header=""):
    current_content, _ = await github_get_file(path)
    if current_content is None:
        current_content = header
    updated = current_content.rstrip() + "\n\n" + new_entry.strip() + "\n"
    return await github_write_file(path, updated, commit_message)


async def compress_memory_if_needed(agent_key):
    """メモリが15エントリを超えたらClaudeで古い部分を圧縮する"""
    folder = KEY_TO_FOLDER.get(agent_key, agent_key)
    path = f"agents/{folder}/memory.md"
    current, _ = await github_get_file(path)
    if not current:
        return
    entries = [e for e in current.split("###") if e.strip()]
    if len(entries) <= 15:
        return
    old_entries = "###".join(entries[:-7])
    recent_entries = "###" + "###".join(entries[-7:])
    try:
        resp = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": f"以下の記憶ログを、重要な決定・方針・合意だけを箇条書き5〜10行で圧縮してください。\n\n{old_entries}"}],
        )
        summary = resp.content[0].text.strip()
        compressed = f"# 個人メモリ\n\n## 長期記憶（圧縮済み）\n{summary}\n\n## 直近の記憶\n{recent_entries.strip()}\n"
        await github_write_file(path, compressed, f"メモリ圧縮 [{agent_key}]")
    except Exception as e:
        print(f"メモリ圧縮エラー ({agent_key}): {e}")


async def append_agent_memory(agent_key, content, timestamp):
    folder = KEY_TO_FOLDER.get(agent_key, agent_key)
    path = f"agents/{folder}/memory.md"
    date_str = timestamp.strftime("%Y-%m-%d %H:%M")
    entry = f"### {date_str}\n{content}"
    success = await github_append_file(path, entry,
        f"記憶更新 [{agent_key}]: {date_str}", header="# 個人メモリ\n")
    await compress_memory_if_needed(agent_key)
    return success


async def append_shared_knowledge(content, timestamp):
    month_str = timestamp.strftime("%Y-%m")
    date_str = timestamp.strftime("%Y-%m-%d %H:%M")
    path = f"knowledge/shared/{month_str}.md"
    entry = f"### {date_str}\n{content}"
    return await github_append_file(path, entry,
        f"共有ナレッジ追加: {date_str}", header=f"# 共有ナレッジ {month_str}\n")


def save_conversation_log(channel_name, messages):
    date_str = datetime.now().strftime("%Y-%m-%d")
    log_dir = os.path.join(BASE_DIR, "logs")
    os.makedirs(log_dir, exist_ok=True)
    filepath = os.path.join(log_dir, f"{date_str}_{channel_name}.md")
    with open(filepath, "a", encoding="utf-8") as f:
        if os.path.getsize(filepath) == 0:
            f.write(f"# 会話ログ {channel_name} {date_str}\n\n")
        for role, text in messages:
            time_str = datetime.now().strftime("%H:%M")
            f.write(f"**[{time_str}] {role}**: {text}\n\n")


async def push_daily_logs():
    date_str = datetime.now().strftime("%Y-%m-%d")
    log_dir = os.path.join(BASE_DIR, "logs")
    if not os.path.exists(log_dir):
        return
    for filename in (f for f in os.listdir(log_dir) if f.startswith(date_str) and f.endswith(".md")):
        filepath = os.path.join(log_dir, filename)
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
        await github_write_file(f"logs/{filename}", content, f"会話ログ: {filename}")
    print(f"ログ日次push完了: {date_str}")


async def webhook_send(webhook_url, agent_key, content):
    agent_name = AGENTS[agent_key]["name"] if agent_key in AGENTS else agent_key
    avatar_url = AGENT_AVATARS.get(agent_key)
    content = strip_emoji(content)
    async with aiohttp.ClientSession() as session:
        webhook = discord.Webhook.from_url(webhook_url, session=session)
        await webhook.send(content=content, username=agent_name, avatar_url=avatar_url)


def save_excel_locally(path, content, timestamp):
    local_path = os.path.join(BASE_DIR, path.replace("/", os.sep))
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    reader = csv.reader(io.StringIO(content.strip()))
    for row_idx, row in enumerate(reader, 1):
        for col_idx, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value.strip())
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    wb.save(local_path)
    return local_path


PROGRESS_KEYWORDS = ["進捗", "状況", "どうなって", "どこまで", "何してる", "何やってる", "報告して", "確認して"]

# ── 構造化ステート管理（state.json）──

STATE_PATH = os.path.join(BASE_DIR, "knowledge", "shared", "state.json")

def load_state():
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"decisions": [], "tasks": {}, "blockers": []}

def save_state(state):
    os.makedirs(os.path.dirname(STATE_PATH), exist_ok=True)
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def state_add_decision(text, agent_key):
    state = load_state()
    state["decisions"].append({
        "text": text, "by": agent_key,
        "ts": datetime.utcnow().isoformat()
    })
    state["decisions"] = state["decisions"][-20:]  # 直近20件
    save_state(state)

def state_update_task(agent_key, status, detail):
    state = load_state()
    state["tasks"][agent_key] = {
        "status": status, "detail": detail,
        "ts": datetime.utcnow().isoformat()
    }
    save_state(state)

def state_add_blocker(agent_key, text):
    state = load_state()
    state["blockers"].append({
        "agent": agent_key, "text": text,
        "ts": datetime.utcnow().isoformat()
    })
    state["blockers"] = state["blockers"][-10:]
    save_state(state)

def get_state_summary():
    """CEOや秘書が読むための現状サマリーを生成"""
    state = load_state()
    lines = []
    if state.get("decisions"):
        lines.append("【決定事項】")
        for d in state["decisions"][-5:]:
            lines.append(f"  {d['by']}: {d['text']}")
    if state.get("tasks"):
        lines.append("【タスク状況】")
        for agent, t in state["tasks"].items():
            lines.append(f"  {agent}: [{t['status']}] {t['detail']}")
    if state.get("blockers"):
        lines.append("【ブロッカー】")
        for b in state["blockers"]:
            lines.append(f"  {b['agent']}: {b['text']}")
    return "\n".join(lines) if lines else "記録なし"


def enrich_content_with_files(content, agent_key):
    """エージェントのメモリに記載されたファイルパスを検出し、内容をメッセージに自動付加する"""
    folder = KEY_TO_FOLDER.get(agent_key, agent_key)
    memory_path = os.path.join(BASE_DIR, "agents", folder, "memory.md")
    if not os.path.exists(memory_path):
        return content
    with open(memory_path, "r", encoding="utf-8") as f:
        memory_text = f.read()
    # メモリ内のファイルパスを抽出（.md / .csv / .txt）
    paths = re.findall(r'projects/[\w/.\-_]+\.(?:md|csv|txt)', memory_text)
    additions = []
    for path in dict.fromkeys(paths):  # 重複除去・順序保持
        local = os.path.join(BASE_DIR, path.replace("/", os.sep))
        if os.path.exists(local):
            try:
                with open(local, "r", encoding="utf-8") as f:
                    file_content = f.read()
                additions.append(f"\n\n[参照ファイル: {path}]\n{file_content[:3000]}\n[/参照ファイル]")
            except Exception:
                pass
    if additions:
        return content + "".join(additions)
    return content


async def collect_progress_report():
    """各エージェントに直接問い合わせて進捗を収集し、秘書が要約して返す"""
    reports = {}
    targets = {
        "sales":     "グラント・カードン",
        "pm":        "グレン・スターンズ",
        "dev":       "スティーブ・ウォズニアック",
        "marketing": "森岡毅",
    }

    for agent_key, agent_name in targets.items():
        agent = AGENTS.get(agent_key)
        if not agent:
            continue
        try:
            resp = claude.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=150,
                system=agent["prompt"],
                messages=[{"role": "user", "content": "菅（秘書）から確認です。現在の進捗・作業状況を3行以内で報告してください。"}],
            )
            reply = strip_emoji(re.sub(r'\[.+?\]', '', resp.content[0].text, flags=re.DOTALL).strip())
            reports[agent_name] = reply
        except Exception as e:
            reports[agent_name] = f"（取得エラー: {e}）"

    # state.jsonのサマリーも追加
    state_summary = get_state_summary()
    report_text = "\n".join(f"【{name}】{text}" for name, text in reports.items())
    if state_summary and state_summary != "記録なし":
        report_text = state_summary + "\n\n" + report_text
    try:
        secretary = AGENTS.get("secretary")
        resp = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            system=secretary["prompt"],
            messages=[{"role": "user", "content": f"以下の各メンバーの進捗報告を受け取りました。CEOへの報告として簡潔にまとめてください。\n\n{report_text}"}],
        )
        summary = strip_emoji(re.sub(r'\[.+?\]', '', resp.content[0].text, flags=re.DOTALL).strip())
    except Exception:
        summary = report_text

    # 秘書のメモリにも保存
    ts = datetime.utcnow()
    await append_agent_memory("secretary", f"[進捗収集結果]\n{report_text}", ts)
    update_agent_prompt("secretary")

    return summary


async def save_file_outputs(reply, timestamp):
    pattern = r'\[FILE:([^\]]+)\]\n?(.*?)\[/FILE\]'
    files = re.findall(pattern, reply, re.DOTALL)
    saved = []
    for path, content in files:
        path = path.strip()
        content = content.strip()
        date_str = timestamp.strftime("%Y-%m-%d %H:%M")
        if path.endswith(".xlsx"):
            local_path = save_excel_locally(path, content, timestamp)
            saved.append(f"Excel: {local_path}")
        else:
            success = await github_write_file(path, content, f"成果物: {path} ({date_str})")
            if success:
                saved.append(path)
    return saved


async def handle_done_ceo(guild, sender_key, summary):
    """[DONE:CEO: サマリー] → 秘書チャンネルにCEOへの完了通知を投稿"""
    sender_name = AGENTS.get(sender_key, {}).get("name", sender_key)
    secretary_ch = None
    for ch in guild.text_channels:
        if "秘書" in ch.name or "secretary" in ch.name.lower():
            secretary_ch = ch
            break
    if not secretary_ch:
        secretary_ch = discord.utils.get(guild.text_channels, name="agent-chat")
    if not secretary_ch:
        return
    ts = datetime.utcnow()
    msg = f"[{sender_name}より完了報告]\n{summary.strip()}"
    webhook_url = AGENT_WEBHOOKS.get("secretary")
    if webhook_url:
        await webhook_send(webhook_url, "secretary", msg)
    else:
        await secretary_ch.send(msg)
    await append_agent_memory("secretary", f"[完了報告/{sender_name}] {summary[:150]}", ts)
    update_agent_prompt("secretary")


async def handle_next_task(guild, sender_key, target_key, task, depth=0):
    """[NEXT:agent: タスク] → target_agentがSonnetでタスクを自律実行。結果に[NEXT:]や[DONE:]があれば連鎖"""
    if depth >= 5:
        return
    agent_chat = discord.utils.get(guild.text_channels, name="agent-chat")
    if not agent_chat:
        return
    target_agent = AGENTS.get(target_key)
    sender_name = AGENTS.get(sender_key, {}).get("name", sender_key)
    if not target_agent:
        return
    target_name = target_agent["name"]
    task = task.strip()

    if WEBHOOK_AGENT_CHAT:
        await webhook_send(WEBHOOK_AGENT_CHAT, sender_key, f"[{target_name}へ引き渡し]\n{task[:100]}")

    async with agent_chat.typing():
        try:
            enriched = enrich_content_with_files(
                f"【{sender_name}からの引き渡しタスク】\n{task}\n\n"
                f"このタスクを実行してください。成果物は[FILE:]形式で出力。"
                f"完了後は [NEXT:agent: 次のタスク] か [DONE:CEO: サマリー] を末尾に追加してください。",
                target_key
            )
            resp = claude.messages.create(
                model=SKILL_MODEL,
                max_tokens=4096,
                system=target_agent["prompt"],
                tools=WEB_SEARCH_TOOL,
                messages=[{"role": "user", "content": enriched}],
            )
            raw = extract_text(resp)
            ts = datetime.utcnow()

            saved_files = await save_file_outputs(raw, ts)

            for sc in re.findall(r'\[SAVE:\s*(.+?)\]', raw, re.DOTALL):
                await append_agent_memory(target_key, sc.strip(), ts)

            clean = re.sub(r'\[FILE:.+?\]\n?.*?\[/FILE\]', '', raw, flags=re.DOTALL)
            clean = re.sub(r'\[(?:SAVE|MSG|NEXT|DONE|SHARE):.+?\]', '', clean, flags=re.DOTALL).strip()
            clean = strip_emoji(clean)
            if clean and WEBHOOK_AGENT_CHAT:
                report = clean
                if saved_files:
                    report += "\n\n保存済み: " + ", ".join(saved_files)
                await webhook_send(WEBHOOK_AGENT_CHAT, target_key, report)

            await append_agent_memory(target_key, f"[{sender_name}からのタスク完了] {task[:80]}", ts)
            await append_agent_memory("secretary", f"[{target_name}がタスク完了] {task[:80]}", ts)
            update_agent_prompt(target_key)
            update_agent_prompt("secretary")

            for next_key, next_task in re.findall(r'\[NEXT:(\w+):\s*(.+?)\]', raw, re.DOTALL):
                next_key = next_key.lower()
                if AGENTS.get(next_key):
                    await handle_next_task(guild, target_key, next_key, next_task, depth + 1)

            for done_summary in re.findall(r'\[DONE:CEO:\s*(.+?)\]', raw, re.DOTALL):
                await handle_done_ceo(guild, target_key, done_summary)

        except Exception as e:
            print(f"handle_next_task エラー ({target_key}): {e}")


async def handle_agent_messaging(guild, sender_key, reply):
    """[SHARE:] [MSG:] [NEXT:] [DONE:] を処理"""
    agent_chat = discord.utils.get(guild.text_channels, name="agent-chat")
    if not agent_chat:
        return
    sender_name = AGENTS.get(sender_key, {}).get("name", sender_key)

    for content in re.findall(r'\[SHARE:\s*(.+?)\]', reply, re.DOTALL):
        content = content.strip()
        if WEBHOOK_AGENT_CHAT:
            await webhook_send(WEBHOOK_AGENT_CHAT, sender_key, f"[全体共有]\n{content}")
        # [SHARE:] の内容を関係エージェントにルーティングして自律実行
        targets = route_message_to_agents(content)
        for target_key in targets:
            if target_key != sender_key and target_key != "secretary":
                await handle_next_task(guild, sender_key, target_key, content)

    for done_summary in re.findall(r'\[DONE:CEO:\s*(.+?)\]', reply, re.DOTALL):
        await handle_done_ceo(guild, sender_key, done_summary)

    for target_key, task in re.findall(r'\[NEXT:(\w+):\s*(.+?)\]', reply, re.DOTALL):
        target_key = target_key.lower()
        if AGENTS.get(target_key):
            await handle_next_task(guild, sender_key, target_key, task)

    for target_key, content in re.findall(r'\[MSG:(\w+):\s*(.+?)\]', reply, re.DOTALL):
        target_key = target_key.lower()
        content = content.strip()
        target_agent = AGENTS.get(target_key)
        if not target_agent:
            continue
        target_name = target_agent["name"]

        if target_key == "secretary":
            ts = datetime.utcnow()
            await append_agent_memory("secretary", f"[{sender_name}より報告] {content}", ts)
            update_agent_prompt("secretary")
            continue

        if WEBHOOK_AGENT_CHAT:
            await webhook_send(WEBHOOK_AGENT_CHAT, sender_key, f"[{target_name}へ]\n{content}")

        async with agent_chat.typing():
            try:
                sender_agent = AGENTS.get(sender_key)
                b_history = []
                last_message = content
                CONCLUSION_WORDS = ["わかった", "了解", "決定", "合意", "それで行こう", "確認した", "やってみる"]

                for round_num in range(2):
                    if round_num % 2 == 0:
                        speaker_key, speaker, listener_name = target_key, target_agent, sender_name
                        b_history.append({"role": "user", "content": f"{sender_name}：{last_message}"})
                    else:
                        speaker_key, speaker, listener_name = sender_key, sender_agent, target_name
                        b_history.append({"role": "user", "content": f"{target_name}：{last_message}"})

                    resp = claude.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=200,
                        system=speaker["prompt"] + f"\n\n今{listener_name}と会話中。簡潔に返答し、[NEXT:] か [DONE:CEO:] で次のアクションを示すこと。",
                        messages=b_history,
                    )
                    raw_reply = resp.content[0].text

                    for nk, nt in re.findall(r'\[NEXT:(\w+):\s*(.+?)\]', raw_reply, re.DOTALL):
                        if AGENTS.get(nk.lower()):
                            await handle_next_task(guild, speaker_key, nk.lower(), nt)
                    for ds in re.findall(r'\[DONE:CEO:\s*(.+?)\]', raw_reply, re.DOTALL):
                        await handle_done_ceo(guild, speaker_key, ds)

                    agent_reply = re.sub(r'\[.+?\]', '', raw_reply, flags=re.DOTALL).strip()
                    agent_reply = strip_emoji(agent_reply)
                    if WEBHOOK_AGENT_CHAT:
                        await webhook_send(WEBHOOK_AGENT_CHAT, speaker_key, agent_reply)

                    b_history.append({"role": "assistant", "content": agent_reply})
                    last_message = agent_reply
                    if any(w in agent_reply for w in CONCLUSION_WORDS) and len(agent_reply) < 120:
                        break

                conv_text = "\n".join(m["content"] for m in b_history)
                ts = datetime.utcnow()
                try:
                    sum_resp = claude.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=150,
                        messages=[{"role": "user", "content": f"以下の会話を1〜2行で要約。決定・合意・次のアクションを中心に。\n\n{conv_text}"}],
                    )
                    summary = sum_resp.content[0].text.strip()
                except Exception:
                    summary = f"{sender_name}と{target_name}の会話：{content[:60]}"

                await append_agent_memory(sender_key, f"[{target_name}との会話] {summary}", ts)
                await append_agent_memory(target_key, f"[{sender_name}との会話] {summary}", ts)
                await append_agent_memory("secretary", f"[{sender_name}x{target_name}] {summary}", ts)
                update_agent_prompt(sender_key)
                update_agent_prompt(target_key)
                update_agent_prompt("secretary")

            except Exception as e:
                print(f"エージェント間会話エラー: {e}")


async def handle_delegation(guild, pm_response):
    delegations = re.findall(r'\[->(\w+):\s*(.+?)\]', pm_response, re.DOTALL)
    if not delegations:
        return
    agent_chat = discord.utils.get(guild.text_channels, name="agent-chat")
    task_board = discord.utils.get(guild.text_channels, name="task-board")
    if not agent_chat:
        return
    for agent_name, task in delegations:
        agent_key = next((k for k in AGENTS if agent_name.lower() in k or k in agent_name.lower()), None)
        if not agent_key:
            continue
        agent = AGENTS[agent_key]
        task = task.strip()
        if task_board:
            await task_board.send(f"[新タスク] {agent['name']}\n{task}")
        if WEBHOOK_AGENT_CHAT:
            await webhook_send(WEBHOOK_AGENT_CHAT, "pm", f"[-> {agent['name']}]\n{task}")
        else:
            await agent_chat.send(f"[PM -> {agent['name']}]\n{task}")
        async with agent_chat.typing():
            try:
                resp = claude.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=512,
                    system=agent["prompt"],
                    messages=[{"role": "user", "content": task}],
                )
                reply = re.sub(r'\[->\w+:.+?\]', '', resp.content[0].text, flags=re.DOTALL).strip()
                reply = strip_emoji(reply)
                if WEBHOOK_AGENT_CHAT:
                    await webhook_send(WEBHOOK_AGENT_CHAT, agent_key, reply)
                else:
                    await agent_chat.send(f"{agent['name']}\n{reply}")
                if task_board:
                    await task_board.send(f"[完了] {agent['name']} / {task[:40]}")
            except Exception as e:
                print(f"委譲エラー: {e}")


async def handle_all_hands(message, content):
    """全体チャンネル: 担当エージェントのみ応答。全エージェントのメモリに即時記録"""
    use_webhook = WEBHOOK_AGENT_CHAT and "agent-chat" in message.channel.name
    primary_agents = route_message_to_agents(content)
    ts = message.created_at
    all_agent_keys = ["secretary", "pm", "sales", "dev", "marketing"]

    for agent_key in primary_agents:
        agent = AGENTS[agent_key]
        async with message.channel.typing():
            try:
                # 全体チャット: 会話履歴なし・短文強制・Haiku固定
                system_allhands = (
                    agent["prompt"]
                    + "\n\n## 全体チャンネルのルール\n"
                    + "返答は3行以内。要点と次のアクションのみ。前置き・まとめ・挨拶不要。"
                )
                resp = claude.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=200,
                    system=system_allhands,
                    messages=[{"role": "user", "content": content}],
                )
                raw_reply = resp.content[0].text

                # [SAVE:] / [NEXT:] / [DONE:] を抽出
                for s in re.findall(r'\[SAVE:\s*(.+?)\]', raw_reply, re.DOTALL):
                    await append_agent_memory(agent_key, s.strip(), ts)

                clean_reply = re.sub(r'\[.+?\]', '', raw_reply, flags=re.DOTALL).strip()
                clean_reply = strip_emoji(clean_reply)

                if use_webhook:
                    await webhook_send(WEBHOOK_AGENT_CHAT, agent_key, clean_reply)
                else:
                    await message.channel.send(f"{agent['name']}\n{clean_reply}")

                # [NEXT:] / [DONE:] パイプライン処理
                if message.guild and any(m in raw_reply for m in ('[NEXT:', '[DONE:')):
                    await handle_agent_messaging(message.guild, agent_key, raw_reply)

                # 返答を全エージェントのメモリに記録（structuredで保存）
                entry = f"[全体会議/{agent['name']}] {clean_reply[:120]}"
                for other_key in all_agent_keys:
                    if other_key != agent_key:
                        await append_agent_memory(other_key, entry, ts)
                        update_agent_prompt(other_key)
                update_agent_prompt(agent_key)

            except Exception as e:
                print(f"全体返答エラー ({agent_key}): {e}")


async def save_standup_log(messages):
    date_str = datetime.now().strftime("%Y-%m-%d")
    path = f"logs/{date_str}_standup.md"
    lines = [f"# 朝会ログ {date_str}\n"]
    for name, text in messages:
        lines.append(f"{name}\n{text}\n")
    await github_write_file(path, "\n".join(lines), f"朝会ログ: {date_str}")


async def run_standup():
    agent_chat = discord.utils.get(bot.get_all_channels(), name="agent-chat")
    if not agent_chat:
        return
    today = datetime.now().strftime("%Y年%m月%d日")
    log = []

    # ── フェーズ1: 朝会（各自の予定共有）──
    opening = f"{today}の朝会を始める。各自、今日の予定・優先タスクを共有してくれ。"
    await webhook_send(WEBHOOK_AGENT_CHAT, "pm", opening)
    log.append(("グレン・スターンズ", opening))

    standup_reports = {}
    for agent_key in ["secretary", "sales", "dev", "marketing"]:
        agent = AGENTS[agent_key]
        try:
            resp = claude.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=150,
                system=agent["prompt"],
                messages=[{"role": "user", "content": f"朝会。今日（{today}）の予定・優先タスクを3行以内で。"}],
            )
            reply = strip_emoji(re.sub(r'\[.+?\]', '', resp.content[0].text, flags=re.DOTALL).strip())
            await webhook_send(WEBHOOK_AGENT_CHAT, agent_key, reply)
            log.append((agent["name"], reply))
            standup_reports[agent_key] = reply
        except Exception as e:
            print(f"朝会エラー ({agent_key}): {e}")

    # ── フェーズ2: PMがメモリ・stateを読んでタスクを割り振る ──
    try:
        state_summary = get_state_summary()
        # 各エージェントのメモリから未完了タスクを収集
        memories = {}
        for ak in ["sales", "marketing", "dev"]:
            folder = KEY_TO_FOLDER.get(ak, ak)
            mem_path = os.path.join(BASE_DIR, "agents", folder, "memory.md")
            if os.path.exists(mem_path):
                with open(mem_path, "r", encoding="utf-8") as f:
                    memories[ak] = f.read()[-1500:]  # 直近1500文字

        reports_text = "\n".join(f"[{k}] {v}" for k, v in standup_reports.items())
        memory_text = "\n".join(f"[{k}メモリ末尾]\n{v}" for k, v in memories.items())

        task_prompt = (
            f"今日（{today}）の朝会報告と各エージェントのメモリを確認した。\n\n"
            f"【朝会報告】\n{reports_text}\n\n"
            f"【現在のステート】\n{state_summary}\n\n"
            f"【各メモリ（末尾）】\n{memory_text}\n\n"
            f"上記を踏まえ、今日中に完了すべき最優先タスクを各エージェントに割り振ってください。\n"
            f"割り振りは [NEXT:agent: タスク内容] 形式で。最大3件まで。\n"
            f"割り振り後、簡潔に「今日の方針」を1行で述べてください。"
        )

        pm_resp = claude.messages.create(
            model=SKILL_MODEL,
            max_tokens=600,
            system=AGENTS["pm"]["prompt"],
            messages=[{"role": "user", "content": task_prompt}],
        )
        pm_raw = pm_resp.content[0].text
        pm_clean = re.sub(r'\[.+?\]', '', pm_raw, flags=re.DOTALL).strip()
        pm_clean = strip_emoji(pm_clean)

        await webhook_send(WEBHOOK_AGENT_CHAT, "pm", pm_clean)
        log.append(("グレン・スターンズ", pm_clean))

        # [NEXT:] を処理して各エージェントが自律実行
        for target_key, task in re.findall(r'\[NEXT:(\w+):\s*(.+?)\]', pm_raw, re.DOTALL):
            target_key = target_key.lower()
            if AGENTS.get(target_key):
                await handle_next_task(agent_chat.guild, "pm", target_key, task)

        # stateにタスク割り振りを記録
        for target_key, task in re.findall(r'\[NEXT:(\w+):\s*(.+?)\]', pm_raw, re.DOTALL):
            state_update_task(target_key.lower(), "assigned", task.strip()[:80])

    except Exception as e:
        print(f"朝会タスク割り振りエラー: {e}")

    await save_standup_log(log)


async def run_sns_drafts():
    task_board = discord.utils.get(bot.get_all_channels(), name="task-board")
    if not task_board:
        return
    today = datetime.now().strftime("%Y年%m月%d日")
    prompt = f"今日（{today}）のThreads投稿を3本作成してください。各500文字以内。ビジネス・マーケティング・ブランド戦略のテーマ。【投稿1】【投稿2】【投稿3】と番号をつけること。"
    try:
        resp = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1200,
            system=AGENTS["marketing"]["prompt"],
            messages=[{"role": "user", "content": prompt}],
        )
        drafts = strip_emoji(resp.content[0].text.strip())
        await task_board.send(f"森岡毅 / Threads投稿案 {datetime.now().strftime('%m/%d')}\n{drafts}")
    except Exception as e:
        print(f"SNS投稿エラー: {e}")


@tasks.loop(time=dtime(0, 0))
async def morning_routine():
    await run_standup()
    await run_sns_drafts()


@tasks.loop(time=dtime(14, 59))
async def nightly_log_push():
    await push_daily_logs()


@bot.event
async def on_ready():
    load_all_agents()
    morning_routine.start()
    nightly_log_push.start()
    print(f"Bot起動完了: {bot.user}")


@bot.event
async def on_message(message):
    if message.author == bot.user:
        return
    if message.author.id != CEO_ID:
        return
    if not AGENTS:
        load_all_agents()

    if message.content == "!ping":
        await message.channel.send("AI会社Bot 稼働中です")
        return
    if message.content == "!reload":
        load_all_agents()
        await message.channel.send("学習ログを再読み込みしました。全エージェントに反映済みです。")
        return
    if message.content == "!standup":
        await message.channel.send("朝会を開始します")
        await run_standup()
        return
    if message.content == "!sns":
        await message.channel.send("森岡さんが投稿案を作成中...")
        await run_sns_drafts()
        return

    # 重複処理防止
    if message.id in processed_message_ids:
        return
    processed_message_ids.add(message.id)
    if len(processed_message_ids) > 500:  # メモリ肥大防止
        processed_message_ids.clear()

    channel_name = message.channel.name if hasattr(message.channel, "name") else ""
    content = message.content.strip()

    if is_all_hands_channel(channel_name):
        if content:
            await handle_all_hands(message, content)
        return

    is_mention = bot.user in message.mentions
    is_agent_channel = any(pattern in channel_name or channel_name in pattern for pattern in CHANNEL_TO_AGENT)

    if is_mention or is_agent_channel:
        content = content.replace(f"<@{bot.user.id}>", "").replace(f"<@!{bot.user.id}>", "").strip()
        if not content:
            await message.channel.send("何かご用件はありますか？")
            return

        agent_key = get_agent_by_channel(channel_name)
        agent = AGENTS[agent_key]

        # 秘書への進捗確認 → 全エージェントに問い合わせて集めて報告
        if agent_key == "secretary" and any(kw in content for kw in PROGRESS_KEYWORDS):
            async with message.channel.typing():
                await message.channel.send("各メンバーに確認します。少々お待ちください。")
                summary = await collect_progress_report()
                webhook_url = AGENT_WEBHOOKS.get("secretary")
                if webhook_url:
                    await webhook_send(webhook_url, "secretary", summary)
                else:
                    await message.channel.send(summary)
            return

        if is_learning_message(content):
            async with message.channel.typing():
                name_to_key = {
                    "森岡": "marketing", "グラント": "sales",
                    "グレン": "pm", "ウォズ": "dev", "ウォズニアック": "dev", "菅": "secretary",
                }
                targeted_key = next((k for name, k in name_to_key.items() if name in content), None)
                if targeted_key:
                    success = await append_agent_memory(targeted_key, content, message.created_at)
                    update_agent_prompt(targeted_key)
                    msg = f"{AGENTS[targeted_key]['name']}のメモリに保存しました" if success else "保存に失敗しました"
                else:
                    success = await append_shared_knowledge(content, message.created_at)
                    msg = "共有ナレッジに保存しました（全エージェント参照）" if success else "保存に失敗しました"
                await message.channel.send(msg)
                try:
                    resp = claude.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=256,
                        system=agent["prompt"],
                        messages=[{"role": "user", "content": content}],
                    )
                    reply = strip_emoji(re.sub(r'\[.+?\]', '', resp.content[0].text, flags=re.DOTALL).strip())
                    webhook_url = AGENT_WEBHOOKS.get(agent_key)
                    if webhook_url:
                        await webhook_send(webhook_url, agent_key, reply)
                    else:
                        await message.channel.send(reply)
                except Exception as e:
                    print(f"学習応答エラー: {e}")
            return

        async with message.channel.typing():
            try:
                ch_id = message.channel.id
                history = conversation_history.get(ch_id, [])
                # メモリ内のファイルパスを自動読み込みしてコンテキストに注入
                enriched = enrich_content_with_files(content, agent_key)
                history.append({"role": "user", "content": enriched})

                model = get_model(content)
                is_skill = model == SKILL_MODEL
                length_rule = (
                    "返答の長さは自由。成果物・分析・リストは必要なだけ詳しく書く。"
                    if is_skill else
                    "返答は5行以内。会話・確認・簡単な質問は2行以内。"
                )
                system_prompt = agent["prompt"] + f"\n\n## 今回の返答スタイル\n{length_rule}"

                kwargs = dict(
                    model=model,
                    max_tokens=4096 if is_skill else 512,
                    system=system_prompt,
                    messages=history,
                )
                if is_skill:
                    kwargs["tools"] = WEB_SEARCH_TOOL

                resp = claude.messages.create(**kwargs)
                reply = extract_text(resp)

                if is_skill:
                    await message.channel.send(f"Sonnet ({SKILL_MODEL}) で実行中...", delete_after=3)

                for save_content in re.findall(r'\[SAVE:\s*(.+?)\]', reply, re.DOTALL):
                    await append_agent_memory(agent_key, save_content.strip(), message.created_at)
                    update_agent_prompt(agent_key)

                saved_files = await save_file_outputs(reply, message.created_at)
                if saved_files:
                    await message.channel.send("GitHub保存完了\n" + "\n".join(f"- {f}" for f in saved_files))

                clean_reply = re.sub(r'\[FILE:.+?\]\n?.*?\[/FILE\]', '', reply, flags=re.DOTALL)
                clean_reply = re.sub(r'\[SAVE:.+?\]', '', clean_reply, flags=re.DOTALL)
                clean_reply = re.sub(r'\[->\w+:.+?\]', '', clean_reply, flags=re.DOTALL).strip()
                clean_reply = strip_emoji(clean_reply)

                history.append({"role": "assistant", "content": clean_reply})
                conversation_history[ch_id] = history[-10:]

                save_conversation_log(channel_name, [("CEO", content), (agent["name"], clean_reply)])

                webhook_url = AGENT_WEBHOOKS.get(agent_key)
                if webhook_url:
                    await webhook_send(webhook_url, agent_key, clean_reply)
                else:
                    await message.channel.send(f"{agent['name']}\n{clean_reply}")

                if agent_key == "pm" and message.guild:
                    await handle_delegation(message.guild, reply)
                if message.guild and any(m in reply for m in ('[SHARE:', '[MSG:', '[NEXT:', '[DONE:')):
                    await handle_agent_messaging(message.guild, agent_key, reply)

            except Exception as e:
                print(f"エラー: {e}")
                await message.channel.send(f"エラーが発生しました: {e}")

    await bot.process_commands(message)


bot.run(TOKEN)
